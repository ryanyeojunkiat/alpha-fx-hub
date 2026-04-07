#!/usr/bin/env python3
"""Create the Alpha FX Hub V6 Callisto trade setup Excel sheet."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── SHEET 1: ORDER SETUP CALCULATOR ──
ws = wb.active
ws.title = "Order Setup"
ws.sheet_properties.tabColor = "FFD700"

# Styles
gold_fill = PatternFill("solid", fgColor="FFD700")
dark_fill = PatternFill("solid", fgColor="1A1A2E")
header_fill = PatternFill("solid", fgColor="16213E")
green_fill = PatternFill("solid", fgColor="0F3D0F")
red_fill = PatternFill("solid", fgColor="3D0F0F")
blue_fill = PatternFill("solid", fgColor="0F1A3D")
input_fill = PatternFill("solid", fgColor="2D2D44")
white_font = Font(name="Arial", color="FFFFFF", size=11)
gold_font = Font(name="Arial", color="FFD700", size=11, bold=True)
gold_font_lg = Font(name="Arial", color="FFD700", size=14, bold=True)
white_bold = Font(name="Arial", color="FFFFFF", size=11, bold=True)
blue_font = Font(name="Arial", color="0000FF", size=11)
green_font = Font(name="Arial", color="00FF00", size=11, bold=True)
red_font = Font(name="Arial", color="FF4444", size=11, bold=True)
thin_border = Border(
    left=Side(style="thin", color="FFD700"),
    right=Side(style="thin", color="FFD700"),
    top=Side(style="thin", color="FFD700"),
    bottom=Side(style="thin", color="FFD700"),
)
center = Alignment(horizontal="center", vertical="center")
left_align = Alignment(horizontal="left", vertical="center")

# Set column widths
widths = {"A": 3, "B": 22, "C": 18, "D": 18, "E": 18, "F": 18, "G": 18, "H": 18}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

# Fill background dark for whole area
for row in range(1, 55):
    for col in range(1, 9):
        cell = ws.cell(row=row, column=col)
        cell.fill = dark_fill
        cell.font = white_font

# ── TITLE ──
ws.merge_cells("B2:G2")
ws["B2"] = "ALPHA FX HUB — V6 CALLISTO ORDER SETUP"
ws["B2"].font = Font(name="Arial", color="FFD700", size=16, bold=True)
ws["B2"].alignment = center

ws.merge_cells("B3:G3")
ws["B3"] = "Gold Engine V6 | Optimized Settings | 0.01 Lot × 5 Orders"
ws["B3"].font = Font(name="Arial", color="AAAAAA", size=10)
ws["B3"].alignment = center

# ── INPUT SECTION ──
ws.merge_cells("B5:C5")
ws["B5"] = "YOUR TRADE INPUT"
ws["B5"].font = gold_font_lg
ws["B5"].fill = header_fill

# Direction
ws["B7"] = "Direction:"
ws["B7"].font = white_bold
ws["C7"] = "BUY"
ws["C7"].font = blue_font
ws["C7"].fill = input_fill
ws["C7"].border = thin_border
ws["C7"].alignment = center
ws["D7"] = '← Type BUY or SELL'
ws["D7"].font = Font(name="Arial", color="888888", size=9, italic=True)

# Entry Price
ws["B8"] = "Entry Price ($):"
ws["B8"].font = white_bold
ws["C8"] = 4500.00
ws["C8"].font = blue_font
ws["C8"].fill = input_fill
ws["C8"].border = thin_border
ws["C8"].number_format = '#,##0.00'
ws["C8"].alignment = center
ws["D8"] = '← Enter your price'
ws["D8"].font = Font(name="Arial", color="888888", size=9, italic=True)

# Capital
ws["B9"] = "Account Balance ($):"
ws["B9"].font = white_bold
ws["C9"] = 1200.00
ws["C9"].font = blue_font
ws["C9"].fill = input_fill
ws["C9"].border = thin_border
ws["C9"].number_format = '#,##0.00'
ws["C9"].alignment = center

# ── FIXED PARAMETERS ──
ws.merge_cells("B11:C11")
ws["B11"] = "OPTIMIZED PARAMETERS"
ws["B11"].font = gold_font_lg
ws["B11"].fill = header_fill

ws["B12"] = "SL (pips):"
ws["B12"].font = white_font
ws["C12"] = 200
ws["C12"].font = gold_font
ws["C12"].alignment = center

ws["B13"] = "Lot per order:"
ws["B13"].font = white_font
ws["C13"] = 0.01
ws["C13"].font = gold_font
ws["C13"].alignment = center

ws["B14"] = "Number of orders:"
ws["B14"].font = white_font
ws["C14"] = 5
ws["C14"].font = gold_font
ws["C14"].alignment = center

ws["B15"] = "Pip value (1 pip = $):"
ws["B15"].font = white_font
ws["C15"] = 0.1
ws["C15"].font = gold_font
ws["C15"].alignment = center

# ── CALCULATED ORDER TABLE ──
ws.merge_cells("B17:H17")
ws["B17"] = "ORDER PLACEMENT TABLE"
ws["B17"].font = gold_font_lg
ws["B17"].fill = header_fill

headers = ["Order", "Lot", "Entry", "SL Price", "TP Price", "TP Pips", "$ if TP Hit"]
for i, h in enumerate(headers):
    cell = ws.cell(row=18, column=i+2)
    cell.value = h
    cell.font = white_bold
    cell.fill = PatternFill("solid", fgColor="2A2A4A")
    cell.alignment = center
    cell.border = thin_border

tp_pips = [200, 400, 600, 800, 1000]
sl_pips = 200

for idx in range(5):
    row = 19 + idx
    # Order #
    ws.cell(row=row, column=2, value=f"#{idx+1}")
    ws.cell(row=row, column=2).font = gold_font
    ws.cell(row=row, column=2).alignment = center
    ws.cell(row=row, column=2).border = thin_border

    # Lot
    ws.cell(row=row, column=3, value=0.01)
    ws.cell(row=row, column=3).font = white_font
    ws.cell(row=row, column=3).alignment = center
    ws.cell(row=row, column=3).border = thin_border

    # Entry = C8
    ws.cell(row=row, column=4).value = f'=C8'
    ws.cell(row=row, column=4).font = white_font
    ws.cell(row=row, column=4).number_format = '#,##0.00'
    ws.cell(row=row, column=4).alignment = center
    ws.cell(row=row, column=4).border = thin_border

    # SL Price: BUY = Entry - SL*pip, SELL = Entry + SL*pip
    ws.cell(row=row, column=5).value = f'=IF(C7="BUY",C8-C12*C15,C8+C12*C15)'
    ws.cell(row=row, column=5).font = red_font
    ws.cell(row=row, column=5).number_format = '#,##0.00'
    ws.cell(row=row, column=5).alignment = center
    ws.cell(row=row, column=5).border = thin_border

    # TP Price: BUY = Entry + TP*pip, SELL = Entry - TP*pip
    ws.cell(row=row, column=6).value = f'=IF(C7="BUY",C8+{tp_pips[idx]}*C15,C8-{tp_pips[idx]}*C15)'
    ws.cell(row=row, column=6).font = green_font
    ws.cell(row=row, column=6).number_format = '#,##0.00'
    ws.cell(row=row, column=6).alignment = center
    ws.cell(row=row, column=6).border = thin_border

    # TP Pips
    ws.cell(row=row, column=7, value=tp_pips[idx])
    ws.cell(row=row, column=7).font = white_font
    ws.cell(row=row, column=7).alignment = center
    ws.cell(row=row, column=7).border = thin_border

    # $ if TP hit
    ws.cell(row=row, column=8).value = f'=C13*{tp_pips[idx]}*10'
    ws.cell(row=row, column=8).font = green_font
    ws.cell(row=row, column=8).number_format = '$#,##0.00'
    ws.cell(row=row, column=8).alignment = center
    ws.cell(row=row, column=8).border = thin_border

# ── RISK SUMMARY ──
ws.merge_cells("B25:H25")
ws["B25"] = "RISK & REWARD SUMMARY"
ws["B25"].font = gold_font_lg
ws["B25"].fill = header_fill

ws["B27"] = "Full SL Loss (worst case):"
ws["B27"].font = white_font
ws.merge_cells("B27:D27")
ws["E27"] = f'=-C13*C12*10*C14'
ws["E27"].font = red_font
ws["E27"].number_format = '$#,##0.00'
ws["E27"].alignment = center

ws["B28"] = "Risk % of Account:"
ws["B28"].font = white_font
ws.merge_cells("B28:D28")
ws["E28"] = '=-E27/C9'
ws["E28"].font = red_font
ws["E28"].number_format = '0.00%'
ws["E28"].alignment = center

ws["B29"] = "TP1 Only + Rest BE:"
ws["B29"].font = white_font
ws.merge_cells("B29:D29")
ws["E29"] = '=H19'
ws["E29"].font = green_font
ws["E29"].number_format = '$#,##0.00'
ws["E29"].alignment = center

ws["B30"] = "All 5 TPs Hit (best case):"
ws["B30"].font = white_font
ws.merge_cells("B30:D30")
ws["E30"] = '=SUM(H19:H23)'
ws["E30"].font = green_font
ws["E30"].number_format = '$#,##0.00'
ws["E30"].alignment = center

ws["B31"] = "Max Daily Loss (2 trades):"
ws["B31"].font = white_font
ws.merge_cells("B31:D31")
ws["E31"] = '=E27*2'
ws["E31"].font = red_font
ws["E31"].number_format = '$#,##0.00'
ws["E31"].alignment = center

ws["B32"] = "Trades to Blow Account:"
ws["B32"].font = white_font
ws.merge_cells("B32:D32")
ws["E32"] = '=INT(C9/ABS(E27))'
ws["E32"].font = Font(name="Arial", color="FFAA00", size=11, bold=True)
ws["E32"].alignment = center

# ── MANAGEMENT RULES ──
ws.merge_cells("B34:H34")
ws["B34"] = "TRADE MANAGEMENT RULES"
ws["B34"].font = gold_font_lg
ws["B34"].fill = header_fill

rules = [
    "1. Place all 5 orders at the SAME entry price",
    "2. All orders share the SAME SL (200 pips from entry)",
    "3. Each order has a DIFFERENT TP (200/400/600/800/1000 pips)",
    "4. When TP1 hits → IMMEDIATELY move SL for #2-#5 to Breakeven + 2 pips",
    "5. Let remaining orders run to their TPs",
    "6. MAX 2 losses per day — then STOP trading",
    "7. London (07:00-16:00 UTC) + NY (12:00-21:00 UTC) sessions ONLY",
    "8. Need Grade B or above from Alpha FX Engine",
    "9. Check MTF alignment (2/3 TFs must agree) before entry",
    "10. CHoCH must be confirmed by BODY close (not wick)",
]
for i, rule in enumerate(rules):
    ws[f"B{36+i}"] = rule
    ws[f"B{36+i}"].font = white_font
    ws.merge_cells(f"B{36+i}:H{36+i}")

# ── BACKTEST STATS ──
ws.merge_cells("B47:H47")
ws["B47"] = "BACKTEST PERFORMANCE (Oct 2025 - Apr 2026)"
ws["B47"].font = gold_font_lg
ws["B47"].fill = header_fill

stats = [
    ("Win Rate:", "51.8%"),
    ("Profit Factor:", "1.47"),
    ("Net P&L:", "+$2,511 (+209%)"),
    ("Max Drawdown:", "33.2%"),
    ("Trades:", "112"),
    ("Profitable Months:", "5 out of 7"),
    ("TP1 Hit Rate:", "50.9%"),
    ("$1,200 → $3,711 in 6 months", ""),
]
for i, (label, val) in enumerate(stats):
    ws[f"B{49+i}"] = label
    ws[f"B{49+i}"].font = white_bold
    ws[f"D{49+i}"] = val
    ws[f"D{49+i}"].font = green_font

# ── SHEET 2: TRADE LOG ──
ws2 = wb.create_sheet("Trade Log")
ws2.sheet_properties.tabColor = "00AA00"

log_headers = ["#", "Date", "Time", "Direction", "Entry Price", "SL Price",
               "TP1", "TP2", "TP3", "TP4", "TP5",
               "Grade", "TPs Hit", "Exit Reason", "P&L ($)", "Balance ($)", "Notes"]

widths2 = [5, 12, 8, 10, 12, 12, 12, 12, 12, 12, 12, 8, 12, 12, 10, 12, 20]
for i, w in enumerate(widths2):
    ws2.column_dimensions[get_column_letter(i+1)].width = w

for row in range(1, 100):
    for col in range(1, 18):
        cell = ws2.cell(row=row, column=col)
        cell.fill = dark_fill
        cell.font = white_font

for i, h in enumerate(log_headers):
    cell = ws2.cell(row=1, column=i+1)
    cell.value = h
    cell.font = white_bold
    cell.fill = header_fill
    cell.alignment = center
    cell.border = thin_border

# Pre-fill formulas for first 50 trades
for row in range(2, 52):
    ws2.cell(row=row, column=1, value=row-1)
    ws2.cell(row=row, column=1).font = Font(name="Arial", color="888888", size=10)
    ws2.cell(row=row, column=1).alignment = center
    # P&L and Balance columns with conditional formatting
    for col in [15, 16]:
        ws2.cell(row=row, column=col).number_format = '$#,##0.00'
        ws2.cell(row=row, column=col).alignment = center

# ── SHEET 3: ENTRY CHECKLIST ──
ws3 = wb.create_sheet("Entry Checklist")
ws3.sheet_properties.tabColor = "FF4444"

ws3.column_dimensions["A"].width = 3
ws3.column_dimensions["B"].width = 8
ws3.column_dimensions["C"].width = 50

for row in range(1, 30):
    for col in range(1, 5):
        cell = ws3.cell(row=row, column=col)
        cell.fill = dark_fill
        cell.font = white_font

ws3.merge_cells("B1:C1")
ws3["B1"] = "PRE-TRADE CHECKLIST"
ws3["B1"].font = gold_font_lg
ws3["B1"].fill = header_fill

checklist = [
    ("TREND", "Daily + H4 trend identified"),
    ("", "2/3 timeframes aligned (H4, H1, M15)"),
    ("REVERSAL", "CHoCH confirmed on M15 (body close)"),
    ("", "Price retesting CHoCH level"),
    ("CONTINUATION", "Candle confirmation (rejection/engulfing)"),
    ("", "SMA44 aligned with direction"),
    ("", "H1 BOS confirms structure"),
    ("SESSION", "London (07-16 UTC) or NY (12-21 UTC) active"),
    ("", "Not within 30min of high-impact news"),
    ("RISK", "Daily losses < 2"),
    ("", "Account has sufficient margin"),
    ("", "Grade B or above from engine"),
    ("EXECUTE", "All 5 orders placed at same entry"),
    ("", "SL set at 200 pips from entry"),
    ("", "TPs set at 200/400/600/800/1000 pips"),
]

for i, (cat, item) in enumerate(checklist):
    row = i + 3
    ws3.cell(row=row, column=2, value=cat)
    ws3.cell(row=row, column=2).font = gold_font if cat else white_font
    ws3.cell(row=row, column=3, value=f"[ ]  {item}")
    ws3.cell(row=row, column=3).font = white_font

out = "/sessions/awesome-nice-johnson/mnt/alpha_fx_hub/AlphaFX_V6_TradeSetup.xlsx"
wb.save(out)
print(f"Saved: {out}")
